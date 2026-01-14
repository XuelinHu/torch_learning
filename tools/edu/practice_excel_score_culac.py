import os
import shutil
import pandas as pd
from openpyxl import load_workbook


def calculate_scores(row):
    """计算 A/B/C + S1~S10（学号去.0 + 等级逻辑）"""
    实训报告原始分 = row['实训报告'] if pd.notna(row.get('实训报告')) else 0.0
    期评成绩 = row['期评成绩'] if pd.notna(row.get('期评成绩')) else 0.0
    班级 = str(row.get('班级', '')).strip() if pd.notna(row.get('班级')) else ""

    # 学号彻底去 .0（兼容浮点/字符串）
    学号_raw = row['学号'] if pd.notna(row.get('学号')) else ""
    学号 = str(学号_raw).strip()
    if 学号.endswith(".0"):
        学号 = 学号[:-2]

    姓名 = str(row.get('姓名', '')).strip() if pd.notna(row.get('姓名')) else ""

    # S7：实训报告满分10分（原始分/10），1位小数，0则空
    S7_val = round(实训报告原始分 / 10, 1)
    S7 = f"{S7_val:.1f}" if S7_val != 0.0 else ""

    # S8：平时表现默认10.0
    S8 = "10.0"

    # S9：总分=期评成绩，1位小数，0则空
    S9_val = round(float(期评成绩), 1) if pd.notna(期评成绩) else 0.0
    S9 = f"{S9_val:.1f}" if S9_val != 0.0 else ""

    # 剩余分（S1~S6 分摊）
    剩余分 = 0.0
    if S9 and S7 and S8:
        剩余分 = round(float(S9) - float(S7) - float(S8), 1)

    def calc_ratio(ratio):
        if 剩余分 == 0.0:
            return ""
        score = round(剩余分 * ratio / 80, 1)
        return f"{score:.1f}" if score != 0.0 else ""

    S1 = calc_ratio(15)
    S2 = calc_ratio(10)
    S3 = calc_ratio(5)
    S4 = calc_ratio(10)
    S5 = calc_ratio(30)

    # S6：补差，确保 S1~S8 总和 = S9
    S6 = ""
    if 剩余分 != 0.0 and all([S1, S2, S3, S4, S5]):
        sum_s1s5 = sum(float(x) for x in [S1, S2, S3, S4, S5])
        S6_val = round(剩余分 - sum_s1s5, 1)
        S6 = f"{S6_val:.1f}" if S6_val != 0.0 else ""

    # S10：等级
    S10 = ""
    if S9:
        s9v = float(S9)
        if s9v >= 90:
            S10 = "优秀"
        elif s9v >= 80:
            S10 = "良好"
        elif s9v >= 70:
            S10 = "中等"
        elif s9v >= 60:
            S10 = "及格"
        else:
            S10 = "不及格"

    return {
        "A": 班级, "B": 学号, "C": 姓名,
        "S1": S1, "S2": S2, "S3": S3, "S4": S4, "S5": S5, "S6": S6,
        "S7": S7, "S8": S8, "S9": S9, "S10": S10
    }


def ensure_headers(ws, needed_headers):
    """确保表头存在；不存在则追加到末尾。返回 header->col_index(1-based) 映射"""
    header_row = 1
    max_col = ws.max_column

    # 读取现有表头
    header_map = {}
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None:
            header_map[str(v).strip()] = c

    # 追加缺失表头
    for h in needed_headers:
        if h not in header_map:
            max_col += 1
            ws.cell(row=header_row, column=max_col).value = h
            header_map[h] = max_col

    return header_map


def write_scores_to_copied_excel(src_excel, dst_excel, sheet_name="Sheet1"):
    """复制Excel -> 在复制件写入 A/B/C + S1~S10"""
    # 1) 复制原始Excel
    shutil.copy2(src_excel, dst_excel)
    print(f"✅ 已复制Excel：\n   源：{src_excel}\n   新：{dst_excel}")

    # 2) 用pandas读取（用于计算）
    df = pd.read_excel(dst_excel, sheet_name=sheet_name)
    print(f"📊 读取成功：{sheet_name}，共 {len(df)} 行数据（不含表头）")

    # 3) openpyxl写回（保留原Excel样式/公式/列宽等）
    wb = load_workbook(dst_excel)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"❌ 工作表不存在：{sheet_name}，实际有：{wb.sheetnames}")
    ws = wb[sheet_name]

    needed = ["A", "B", "C", "S1", "S2", "S3", "S4", "S5", "S6", "S7", "S8", "S9", "S10"]
    header_map = ensure_headers(ws, needed)

    # 4) 逐行写入（Excel第2行对应df第0行）
    start_excel_row = 2
    for i, row in df.iterrows():
        data = calculate_scores(row)
        excel_row = start_excel_row + i
        for key, value in data.items():
            col = header_map[key]
            ws.cell(row=excel_row, column=col).value = value

    wb.save(dst_excel)
    print("🎉 写入完成：已将 A/B/C + S1~S10 写入复制后的Excel。")


def main():
    # === 你的原始Excel ===
    src_excel = r"D:\edu_lt\教案\企业网设计与管理实训-胡学林\作业成绩汇总-2025\24企信30班_原始版.xlsx"

    # === 生成一个“复制件” ===
    dst_excel = r"D:\edu_lt\教案\企业网设计与管理实训-胡学林\作业成绩汇总-2025\24企信30班_写入S1-S10.xlsx"

    if not os.path.exists(src_excel):
        print(f"❌ 源Excel不存在：{src_excel}")
        return

    write_scores_to_copied_excel(src_excel, dst_excel, sheet_name="Sheet1")
    print(f"📌 输出文件：{dst_excel}")


if __name__ == "__main__":
    main()
